import streamlit as st
import pandas as pd
import os
import zipfile
import shutil
from datetime import datetime
from io import BytesIO
import chardet
from openpyxl import load_workbook
import tempfile
import win32com.client
import pythoncom

st.set_page_config(page_title='Excel分割合併工具', page_icon='📝')

st.markdown("""
    <style>
    div.block-container {
        padding-top: 3.5rem;
    }
    </style>
    """, unsafe_allow_html=True)
    
def detect_encoding(file):
    raw_data = file.read(10000)
    result = chardet.detect(raw_data)
    return result['encoding']

def read_uploaded_file(uploaded_file, header_rows=1):
    if uploaded_file.name.endswith(('.xlsx', '.xls')):
        df = pd.read_excel(uploaded_file, dtype={'選項ID': str})
    elif uploaded_file.name.endswith('.csv'):
        uploaded_file.seek(0)
        encoding = detect_encoding(uploaded_file)
        uploaded_file.seek(0)
        df = pd.read_csv(uploaded_file, encoding=encoding, dtype={'選項ID': str})
    else:
        df = None
    if df is not None:
        df = df.iloc[header_rows:]
    return df

def unprotect_excel_sheet(file_path):
    pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        # 嘗試打開 Excel 檔案，並解除保護
        workbook = excel.Workbooks.Open(os.path.abspath(file_path))
        for sheet in workbook.Sheets:
            sheet.Unprotect()
        # 儲存更改並關閉檔案
        workbook.Save()
        workbook.Close(SaveChanges=True)
        print(f"{file_path} 解除保護成功")
    except Exception as e:
        print(f"處理 {file_path} 時發生錯誤: {e}")
    finally:
        # 確保正確關閉工作簿與 Excel 應用程式
        excel.Quit()
        pythoncom.CoUninitialize()

def unzip_and_unprotect(zip_file_path, extract_to, progress_bar, status_text):
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)

    unprotected_files = []
    files = [f for f in os.listdir(extract_to) if f.endswith('.xlsx')]
    total_files = len(files)

    for idx, file in enumerate(files):
        file_path = os.path.join(extract_to, file)
        print(f"正在處理: {file_path}")
        try:
            unprotect_excel_sheet(file_path)
            unprotected_files.append(file_path)
        except Exception as e:
            print(f"在處理 {file_path} 時發生錯誤，跳過此檔案: {e}")

        progress = (idx + 1) / total_files
        progress_bar.progress(progress)
        status_text.text(f"解除保護進度: {idx + 1} / {total_files}")
    
    return unprotected_files

def split_by_unique_ids(df, split_column, split_size):
    unique_ids = df[split_column].unique()
    id_map = {ID: idx + 1 for idx, ID in enumerate(unique_ids)}
    df['編號'] = df[split_column].map(id_map)
    max_id = max(id_map.values())
    chunks = []
    log_details = []
    for num_start in range(1, max_id + 1, split_size):
        num_end = num_start + split_size - 1
        chunk = df[(df['編號'] >= num_start) & (df['編號'] <= num_end)].copy()
        if not chunk.empty:
            chunk = chunk.drop(columns=['編號'])
            chunks.append(chunk)
            num_rows = len(chunk)
            num_ids = chunk[split_column].nunique()
            log_details.append(f'包含 {num_rows} 筆資料，涵蓋 {num_ids} 組{split_column}')
    return chunks, log_details

def split_by_row_count(df, split_column, split_size):
    chunks = []
    current_chunk = []
    current_chunk_size = 0
    log_details = []
    grouped = df.groupby(split_column)
    for group, data in grouped:
        if current_chunk_size + len(data) <= split_size:
            current_chunk.append(data)
            current_chunk_size += len(data)
        else:
            chunks.append(pd.concat(current_chunk))
            current_chunk = [data]
            current_chunk_size = len(data)
    if current_chunk:
        chunks.append(pd.concat(current_chunk))
    for i, chunk in enumerate(chunks):
        num_rows = len(chunk)
        log_details.append(f'包含 {num_rows} 筆資料')
    return chunks, log_details

def zip_output_directory(output_dir):
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for foldername, subfolders, filenames in os.walk(output_dir):
            for filename in filenames:
                file_path = os.path.join(foldername, filename)
                zip_file.write(file_path, os.path.relpath(file_path, output_dir))
    zip_buffer.seek(0)
    return zip_buffer

def reset_file_uploader():
    st.session_state['split_file_uploader_key'] += 1
    st.session_state['merge_file_uploader_key'] += 1  

def main():
    if 'split_file_uploader_key' not in st.session_state:
        st.session_state['split_file_uploader_key'] = 0
    if 'merge_file_uploader_key' not in st.session_state:
        st.session_state['merge_file_uploader_key'] = 0

    tab1, tab2 = st.tabs(["Excel分割", "Excel合併"])

    with tab1:
        st.header("91APP EXCEL分割工具")
        st.write("\n")
        col1, col2 = st.columns(2)

        with col1:
            split_logic = st.selectbox("選擇分割邏輯 (同組資料均不拆散)", ["計算不重覆ID數 (Shopee模式)", "計算表格行數 (Yahoo模式)"])
            split_column_default = "商品 ID" if split_logic == "計算不重覆ID數 (Shopee模式)" else "賣場編號"
            split_column = st.text_input("分割依據的欄位標題", value=split_column_default)

        with col2:
            header_rows = st.number_input("標題佔幾行?", min_value=0, value=1, key="split")
            split_size_label = '各檔案的不重覆ID數' if split_logic == '計算不重覆ID數 (Shopee模式)' else '各檔案的行數上限'
            split_size = st.number_input(split_size_label, min_value=1, value=1000, key="split_size")

        uploaded_file = st.file_uploader("上傳 CSV 或 EXCEL", type=['csv', 'xlsx'],
                                         key='split_file_uploader_' + str(st.session_state['split_file_uploader_key']))

        if uploaded_file is not None:
            start_split = st.button("開始分割")
            if start_split:
                st.write("\n")
                progress_bar = st.progress(0)
                status_text = st.empty()
                status_text.text("開始處理文件，請稍後...")
                
                df = read_uploaded_file(uploaded_file, header_rows=header_rows)
                if df is None:
                    st.error("無法讀取上傳的檔案，請確認檔案格式是否正確。")
                    return             
                if split_column not in df.columns:
                    st.error(f"上傳的檔案中找不到 {split_column} 欄位，請重新確認。")
                    return
                    
                total_rows = len(df)
                today_date = datetime.now().strftime('%Y%m%d')
                output_dir = os.path.join(tempfile.gettempdir(), today_date)

                if os.path.exists(output_dir):
                    shutil.rmtree(output_dir)
                os.makedirs(output_dir, exist_ok=True)

                if split_logic == "計算不重覆ID數 (Shopee模式)":
                    chunks, log_details = split_by_unique_ids(df, split_column, split_size)
                else:
                    chunks, log_details = split_by_row_count(df, split_column, split_size)

                total_chunks = len(chunks)
                
                for idx, chunk in enumerate(chunks):
                    output_path = os.path.join(output_dir, f'{idx + 1}.xlsx')
                    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
                        chunk.to_excel(writer, index=False)
                        worksheet = writer.sheets['Sheet1']
                        for col_num, value in enumerate(chunk.columns.values):
                            worksheet.write(0, col_num, value)
                        if '選項ID' in chunk.columns:
                            format1 = writer.book.add_format({'num_format': '@'})
                            col_idx = chunk.columns.get_loc('選項ID')
                            worksheet.set_column(col_idx, col_idx, None, format1)
                    progress_fraction = (idx + 1) / total_chunks
                    progress_bar.progress(progress_fraction)
                    status_text.text(f"處理進度: {idx + 1} / {total_chunks}")

                progress_bar.empty()
                status_text.empty()

                log_file_path = os.path.join(output_dir, f'{today_date}_分割log.txt')
                with open(log_file_path, 'w') as log_file:
                    log_file.write(f"總共處理了 {total_rows} 筆資料，分成 {total_chunks} 份檔案\n")
                    for idx, log in enumerate(log_details):
                        log_file.write(f'保存文件: {idx + 1}.xlsx，{log}\n')

                zip_buffer = zip_output_directory(output_dir)
                shutil.rmtree(output_dir)

                st.success("檔案處理完成！")

                st.download_button(
                    label="下載分割檔案",
                    data=zip_buffer,
                    file_name=f'{today_date}_分割.zip',
                    mime='application/zip',
                    on_click=reset_file_uploader
                )

    with tab2:
        st.header("91APP EXCEL合併工具")
        st.write("\n")
        header_rows = st.number_input("標題佔幾行?", min_value=0, value=6, key="merge")
        uploaded_file = st.file_uploader("上傳包含 CSV 或 EXCEL 的 ZIP 檔", type=['zip'],
                                         key='merge_file_uploader_' + str(st.session_state['merge_file_uploader_key']))

        if uploaded_file is not None:
            start_merge = st.button("開始合併")
            if start_merge:
                st.write("\n")
                        
                temp_dir = tempfile.mkdtemp()

                # 先顯示進度條與狀態文字
                progress_bar = st.progress(0)
                status_text = st.empty()
                status_text.text("正在解壓縮並解除工作表保護，請稍候...")

                # 暫存 ZIP 檔案並解壓縮和解除保護
                with tempfile.NamedTemporaryFile(delete=False) as temp_zip:
                    temp_zip.write(uploaded_file.read())
                    temp_zip.close()

                    unprotected_files = unzip_and_unprotect(temp_zip.name, temp_dir, progress_bar, status_text)

                # 取消保護完成後，重設進度條和狀態
                progress_bar.empty()
                status_text.empty()

                # 重設進度條為 0，並開始合併
                total_files = len(unprotected_files)
                progress_bar = st.progress(0)
                status_text = st.empty()
                merged_data = pd.DataFrame()

                # 合併檔案
                header_workbook = None
                log_details = []
                today_date = datetime.now().strftime('%Y%m%d')

                for idx, file_path in enumerate(unprotected_files):
                    try:
                        workbook = load_workbook(file_path)
                        sheet = workbook.active

                        if header_workbook is None:
                            header_workbook = load_workbook(file_path)
                            merged_workbook = load_workbook(file_path)
                            merged_sheet = merged_workbook.active

                        for row in sheet.iter_rows(min_row=header_rows + 1, values_only=True):
                            merged_sheet.append(row)

                        log_details.append(f"成功處理 Excel 檔案: {os.path.basename(file_path)}")
                    except Exception as e:
                        log_details.append(f"無法讀取 Excel 檔案 {os.path.basename(file_path)}，錯誤訊息: {e}")

                    progress = (idx + 1) / total_files
                    progress_bar.progress(progress)
                    status_text.text(f"處理進度: {idx + 1} / {total_files}")

                progress_bar.empty()
                status_text.empty()

                # 完成合併，提供下載
                with st.spinner('就快完成了...'):
                    output = BytesIO()
                    if 'merged_workbook' in locals():
                        merged_workbook.save(output)
                    output.seek(0)

                    log_output = BytesIO()
                    log_content = '\n'.join(log_details)
                    log_output.write(log_content.encode('utf-8'))
                    log_output.seek(0)

                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        zip_file.writestr(f'{today_date}_合併.xlsx', output.getvalue())
                        zip_file.writestr(f'{today_date}_合併log.txt', log_output.getvalue())

                    zip_buffer.seek(0)

                shutil.rmtree(temp_dir)

                st.success("檔案處理完成！")
                
                st.download_button(
                    label="下載合併檔案",
                    data=zip_buffer,
                    file_name=f'{today_date}_合併.zip',
                    mime='application/zip'
                )

                
if __name__ == '__main__':
    main()
